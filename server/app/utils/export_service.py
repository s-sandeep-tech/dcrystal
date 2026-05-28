import os
import logging
from datetime import datetime, timezone, timedelta
from decimal import Decimal

logger = logging.getLogger(__name__)

# IST timezone offset
IST = timezone(timedelta(hours=5, minutes=30))

# Directory where exports are saved
if os.path.isdir('/app/uploads'):
    EXPORTS_DIR = '/app/uploads/exports'
else:
    # Fallback for local development outside docker
    EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'uploads', 'exports')


def _ensure_exports_dir():
    os.makedirs(EXPORTS_DIR, exist_ok=True)


def _safe_float(val):
    try:
        if val is None:
            return 0.0
        return float(val)
    except Exception:
        return 0.0


def _safe_int(val):
    try:
        if val is None:
            return 0
        return int(float(val))
    except Exception:
        return 0


def generate_outstanding_po_export(filters: dict) -> str:
    """
    Query OutstandingPurchaseOrderStatusSnapshot with the given filters,
    write a formatted .xlsx file to EXPORTS_DIR, and return the filename.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(f"openpyxl not installed: {e}")

    from app.extensions import db
    from app.models import OutstandingPurchaseOrderStatusSnapshot
    from sqlalchemy import func

    _ensure_exports_dir()

    # ── 1. Build Query ────────────────────────────────────────────────────────
    search              = filters.get('search', '').strip()
    classification_owner = filters.get('classification_owner', '')
    make_owner          = filters.get('make_owner', '')
    collection_owner    = filters.get('collection_owner', '')
    purchase_ro         = filters.get('purchase_ro', '')
    party               = filters.get('party', '')
    classification      = filters.get('classification', '')
    make                = filters.get('make', '')
    collection          = filters.get('collection', '')
    section             = filters.get('section', '')
    division            = filters.get('division', '')
    group               = filters.get('group', '')
    purity              = filters.get('purity', '')
    age_min             = filters.get('age_min')
    age_max             = filters.get('age_max')
    exclude_receipt     = filters.get('exclude_receipt', False)

    if isinstance(age_min, str) and age_min.strip():
        try: age_min = int(age_min)
        except Exception: age_min = None
    if isinstance(age_max, str) and age_max.strip():
        try: age_max = int(age_max)
        except Exception: age_max = None
    if isinstance(exclude_receipt, str):
        exclude_receipt = exclude_receipt.lower() == 'true'

    q = db.session.query(OutstandingPurchaseOrderStatusSnapshot)
    if search:
        q = q.filter(
            OutstandingPurchaseOrderStatusSnapshot.classification_owner.ilike(f'%{search}%') |
            OutstandingPurchaseOrderStatusSnapshot.make_owner.ilike(f'%{search}%') |
            OutstandingPurchaseOrderStatusSnapshot.collection_owner.ilike(f'%{search}%') |
            OutstandingPurchaseOrderStatusSnapshot.party.ilike(f'%{search}%') |
            OutstandingPurchaseOrderStatusSnapshot.order_number.ilike(f'%{search}%')
        )
    if classification_owner:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.classification_owner == classification_owner)
    if make_owner:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.make_owner == make_owner)
    if collection_owner:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.collection_owner == collection_owner)
    if purchase_ro:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.purchase_ro == purchase_ro)
    if party:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.party == party)
    if classification:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.classification == classification)
    if make:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.make == make)
    if collection:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.collection == collection)
    if section:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.section == section)
    if division:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.division == division)
    if group:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.group == group)
    if purity:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.purity == purity)
    if age_min is not None:
        q = q.filter(func.current_date() - OutstandingPurchaseOrderStatusSnapshot.order_date >= age_min)
    if age_max is not None:
        q = q.filter(func.current_date() - OutstandingPurchaseOrderStatusSnapshot.order_date <= age_max)
    if exclude_receipt:
        q = q.filter(OutstandingPurchaseOrderStatusSnapshot.receipt_present != 'Y')

    rows = q.order_by(
        OutstandingPurchaseOrderStatusSnapshot.classification_owner,
        OutstandingPurchaseOrderStatusSnapshot.make_owner,
        OutstandingPurchaseOrderStatusSnapshot.collection_owner,
        OutstandingPurchaseOrderStatusSnapshot.order_date
    ).all()

    # ── 2. Build Active Filter Summary ────────────────────────────────────────
    active_filters = []
    if search:              active_filters.append(('Search', search))
    if classification_owner: active_filters.append(('Classification Owner', classification_owner))
    if make_owner:          active_filters.append(('Make Owner', make_owner))
    if collection_owner:    active_filters.append(('Collection Owner', collection_owner))
    if purchase_ro:         active_filters.append(('Purchase RO', purchase_ro))
    if party:               active_filters.append(('Party', party))
    if classification:      active_filters.append(('Classification', classification))
    if make:                active_filters.append(('Make', make))
    if collection:          active_filters.append(('Collection', collection))
    if section:             active_filters.append(('Section', section))
    if division:            active_filters.append(('Division', division))
    if group:               active_filters.append(('Group', group))
    if purity:              active_filters.append(('Purity', purity))
    if age_min is not None: active_filters.append(('Min Age (Days)', str(age_min)))
    if age_max is not None: active_filters.append(('Max Age (Days)', str(age_max)))
    if exclude_receipt:     active_filters.append(('Exclude Receipt', 'Yes'))
    if not active_filters:
        active_filters.append(('Filters', 'None — showing all data'))

    # ── 3. Create Workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Outstanding Purchase Orders'

    # ── Colour Palette ────────────────────────────────────────────────────────
    PRIMARY_FILL   = PatternFill('solid', fgColor='1E3A5F')   # Dark navy
    HEADER_FILL    = PatternFill('solid', fgColor='2563EB')   # Blue
    FILTER_FILL    = PatternFill('solid', fgColor='EFF6FF')   # Light blue tint
    FOOTER_FILL    = PatternFill('solid', fgColor='F1F5F9')   # Light gray
    ALT_FILL       = PatternFill('solid', fgColor='F8FAFC')   # Very light gray for alt rows

    TITLE_FONT     = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    SUB_FONT       = Font(name='Calibri', size=10, color='94A3B8', italic=True)
    HEADER_FONT    = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    FILTER_KEY_F   = Font(name='Calibri', bold=True, size=9, color='1E40AF')
    FILTER_VAL_F   = Font(name='Calibri', size=9, color='1E3A5F')
    DATA_FONT      = Font(name='Calibri', size=9, color='1E293B')
    FOOTER_FONT    = Font(name='Calibri', bold=True, size=9, color='1E293B')
    SECTION_FONT   = Font(name='Calibri', bold=True, size=9, color='64748B')

    THIN_BORDER = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Number formats
    NUM_FMT_3DP = '#,##0.000'
    NUM_FMT_INT = '#,##0'
    NUM_FMT_DATE = 'DD-MMM-YYYY'

    # Total columns in the sheet
    TOTAL_COLS = 20

    # ── Row 1: Report Title ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1,
                         value='Outstanding Purchase Order Status Report')
    title_cell.font = TITLE_FONT
    title_cell.fill = PRIMARY_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Row 2: Generated timestamp ────────────────────────────────────────────
    now_ist = datetime.now(IST)
    gen_time_str = now_ist.strftime('%d %b %Y, %I:%M %p IST')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    gen_cell = ws.cell(row=2, column=1, value=f'Generated: {gen_time_str}   |   Total Records: {len(rows)}')
    gen_cell.font = Font(name='Calibri', size=9, color='64748B')
    gen_cell.fill = PatternFill('solid', fgColor='F8FAFC')
    gen_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Row 3: Blank separator ────────────────────────────────────────────────
    cur_row = 3

    # ── Filter Criteria Section ───────────────────────────────────────────────
    # Section header
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=TOTAL_COLS)
    sec_cell = ws.cell(row=cur_row, column=1, value='  APPLIED FILTERS')
    sec_cell.font = SECTION_FONT
    sec_cell.fill = PatternFill('solid', fgColor='E2E8F0')
    sec_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[cur_row].height = 16
    cur_row += 1

    # Filter rows: 2 filters per row (col 1-2, col 3-4, etc.)
    for i in range(0, len(active_filters), 3):
        for j in range(3):
            if i + j < len(active_filters):
                label, value = active_filters[i + j]
                base = j * 4 + 1  # col 1, 5, 9
                lbl_cell = ws.cell(row=cur_row, column=base, value=label)
                lbl_cell.font = FILTER_KEY_F
                lbl_cell.fill = FILTER_FILL
                lbl_cell.alignment = Alignment(horizontal='right', vertical='center')
                lbl_cell.border = THIN_BORDER
                ws.merge_cells(start_row=cur_row, start_column=base, end_row=cur_row, end_column=base)

                val_cell = ws.cell(row=cur_row, column=base + 1, value=value)
                val_cell.font = FILTER_VAL_F
                val_cell.fill = FILTER_FILL
                val_cell.alignment = Alignment(horizontal='left', vertical='center')
                val_cell.border = THIN_BORDER
                ws.merge_cells(start_row=cur_row, start_column=base + 1, end_row=cur_row, end_column=base + 3)
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

    # Blank separator before data
    cur_row += 1

    # ── Data Headers ──────────────────────────────────────────────────────────
    COLUMNS = [
        # (header_label, column_key, width, number_format, alignment)
        ('Order Number',          'order_number',          20, None,         'left'),
        ('Order Date',            'order_date',            14, NUM_FMT_DATE, 'center'),
        ('Age (Days)',            'age_days',              10, NUM_FMT_INT,  'right'),
        ('Party',                 'party',                 25, None,         'left'),
        ('Classification Owner',  'classification_owner',  22, None,         'left'),
        ('Make Owner',            'make_owner',            22, None,         'left'),
        ('Collection Owner',      'collection_owner',      22, None,         'left'),
        ('Purchase RO',           'purchase_ro',           16, None,         'left'),
        ('Classification',        'classification',        18, None,         'left'),
        ('Make',                  'make',                  18, None,         'left'),
        ('Collection',            'collection',            18, None,         'left'),
        ('Section',               'section',               14, None,         'left'),
        ('Division',              'division',              14, None,         'left'),
        ('Group',                 'group',                 14, None,         'left'),
        ('Purity',                'purity',                12, None,         'center'),
        ('Receipt',               'receipt_present',       10, None,         'center'),
        ('Order Pcs',             'order_pieces',          12, NUM_FMT_INT,  'right'),
        ('Order Wt (g)',          'order_weight',          14, NUM_FMT_3DP,  'right'),
        ('Accepted Pcs',          'accepted_pieces',       12, NUM_FMT_INT,  'right'),
        ('Accepted Wt (g)',       'accepted_weight',       14, NUM_FMT_3DP,  'right'),
    ]

    header_row = cur_row
    for col_idx, (label, _, width, _, align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30

    # Enable auto-filter on data
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(rows)}"
    )
    ws.freeze_panes = f'A{header_row + 1}'

    # ── Data Rows ─────────────────────────────────────────────────────────────
    today = datetime.now(IST).date()

    # Totals accumulators
    total_order_pcs = 0
    total_order_wt = 0.0
    total_accepted_pcs = 0
    total_accepted_wt = 0.0

    for row_idx, record in enumerate(rows, start=1):
        xls_row = header_row + row_idx
        alt = (row_idx % 2 == 0)
        row_fill = ALT_FILL if alt else None

        # Pre-compute age
        if record.order_date:
            age_days = (today - record.order_date).days
        else:
            age_days = None

        order_pcs = _safe_int(record.order_pieces)
        order_wt  = _safe_float(record.order_weight)
        acc_pcs   = _safe_int(record.accepted_pieces)
        acc_wt    = _safe_float(record.accepted_weight)

        total_order_pcs   += order_pcs
        total_order_wt    += order_wt
        total_accepted_pcs += acc_pcs
        total_accepted_wt  += acc_wt

        row_values = [
            record.order_number or '',
            record.order_date,           # date object → openpyxl will format it
            age_days,
            record.party or '',
            record.classification_owner or '',
            record.make_owner or '',
            record.collection_owner or '',
            record.purchase_ro or '',
            record.classification or '',
            record.make or '',
            record.collection or '',
            record.section or '',
            record.division or '',
            record.group or '',
            record.purity or '',
            record.receipt_present or '',
            order_pcs,
            order_wt,
            acc_pcs,
            acc_wt,
        ]

        for col_idx, (_, _, _, num_fmt, align) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=xls_row, column=col_idx, value=row_values[col_idx - 1])
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal=align, vertical='center')
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if num_fmt:
                cell.number_format = num_fmt

        ws.row_dimensions[xls_row].height = 15

    # ── Footer Totals Row ─────────────────────────────────────────────────────
    footer_row = header_row + len(rows) + 1
    footer_values = {
        1: 'TOTALS',
        17: total_order_pcs,
        18: total_order_wt,
        19: total_accepted_pcs,
        20: total_accepted_wt,
    }
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=footer_row, column=col_idx,
                       value=footer_values.get(col_idx, ''))
        cell.font = FOOTER_FONT
        cell.fill = FOOTER_FILL
        cell.border = THIN_BORDER
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif col_idx in (17, 19):
            cell.number_format = NUM_FMT_INT
            cell.alignment = Alignment(horizontal='right', vertical='center')
        elif col_idx in (18, 20):
            cell.number_format = NUM_FMT_3DP
            cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[footer_row].height = 18

    # ── Save ──────────────────────────────────────────────────────────────────
    timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f'outstanding_po_export_{timestamp}.xlsx'
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    logger.info(f'Export saved: {filepath}  ({len(rows)} records)')
    return filename


def generate_pending_acceptance_export(filters: dict) -> str:
    """
    Query PendingAcceptanceAction with filters, explode action_data (specifically CANCEL actions),
    and write to an .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(f"openpyxl not installed: {e}")

    from app.extensions import db
    from app.models.snapshots import PendingAcceptanceAction
    from sqlalchemy import func

    _ensure_exports_dir()

    # ── 1. Build Query ────────────────────────────────────────────────────────
    status_filter = filters.get('status_filter', 'pending_to_deliver_not_barcoded')
    collection_owner = filters.get('collection_owner', '')
    make_owner = filters.get('make_owner', '')
    supplier = filters.get('supplier', '')
    collection = filters.get('collection', '')
    action_type = filters.get('action_action_type', '')  # Optional: filter by action type (e.g. CANCEL)

    q = db.session.query(PendingAcceptanceAction)
    
    if status_filter:
        q = q.filter(PendingAcceptanceAction.status_filter == status_filter)
    if collection_owner:
        q = q.filter(PendingAcceptanceAction.collection_owner == collection_owner)
    if make_owner:
        q = q.filter(PendingAcceptanceAction.make_owner == make_owner)
    if supplier:
        q = q.filter(PendingAcceptanceAction.supplier == supplier)
    if collection:
        q = q.filter(PendingAcceptanceAction.collection == collection)
    if action_type:
        q = q.filter(PendingAcceptanceAction.action_type == action_type)

    actions = q.order_by(PendingAcceptanceAction.created_at.desc()).all()

    # ── 2. Explode Data ───────────────────────────────────────────────────────
    # We flatten the action_data (JSON) into individual rows per PO
    po_rows = []
    for action in actions:
        data = action.action_data
        if not data:
            continue
            
        items_to_process = []
        schedules_str = ""
        
        if isinstance(data, list):
            # Old format or simple CANCEL action
            if action.action_type == 'CONTINUE':
                # Legacy CONTINUE format was a list of schedules WITHOUT PO metadata
                schedules_str = ", ".join([f"{s.get('weight')} ({s.get('delivery_date')})" if s.get('delivery_date') else str(s.get('weight')) for s in data])
                total_wt = sum([float(s.get('weight') or 0) for s in data])
                # Add a summary row for legacy data
                po_rows.append({
                    'action_id': action.id,
                    'action_type': action.action_type,
                    'action_date': action.created_at.strftime('%Y-%m-%d %H:%M') if action.created_at else '',
                    'username': action.username or '',
                    'reason': action.reason or '',
                    'collection_owner': action.collection_owner or '',
                    'make_owner': action.make_owner or '',
                    'supplier': action.supplier or '',
                    'collection': action.collection or '',
                    'status_filter': action.status_filter or '',
                    'po_number': 'GROUP_RESCHEDULE',
                    'po_date': 'N/A',
                    'order_weight': total_wt,
                    'order_piece': 0,
                    'schedules': schedules_str
                })
            else:
                items_to_process = data
        elif isinstance(data, dict):
            # New structure with schedules and unselected_pos for CONTINUE
            if action.action_type == 'CONTINUE':
                items_to_process = data.get('unselected_pos', [])
                schedules = data.get('schedules', [])
                # Format: "Weight (Date), Weight (Date)..."
                schedules_str = ", ".join([f"{s.get('weight')} ({s.get('delivery_date')})" if s.get('delivery_date') else str(s.get('weight')) for s in schedules])
                
                if schedules:
                    total_wt = sum([float(s.get('weight') or 0) for s in schedules])
                    # Add a summary row for the rescheduled items
                    po_rows.append({
                        'action_id': action.id,
                        'action_type': 'CONTINUE (GROUP_RESCHEDULE)',
                        'action_date': action.created_at.strftime('%Y-%m-%d %H:%M') if action.created_at else '',
                        'username': action.username or '',
                        'reason': action.reason or '',
                        'collection_owner': action.collection_owner or '',
                        'make_owner': action.make_owner or '',
                        'supplier': action.supplier or '',
                        'collection': action.collection or '',
                        'status_filter': action.status_filter or '',
                        'po_number': 'GROUP_RESCHEDULE',
                        'po_date': 'N/A',
                        'order_weight': total_wt,
                        'order_piece': 0,
                        'schedules': schedules_str
                    })
            elif action.action_type == 'CANCEL':
                items_to_process = data.get('selected', []) + data.get('unselected', [])

        for item in items_to_process:
            # item is a dict like {'po_number': '...', 'po_date': '...', 'total_weight': ..., 'order_piece': ..., 'vendor': '...'}
            po_num = item.get('po_number') or item.get('order_number')
            if not po_num:
                continue
                
            po_rows.append({
                'action_id': action.id,
                'action_type': action.action_type,
                'action_date': action.created_at.strftime('%Y-%m-%d %H:%M') if action.created_at else '',
                'username': action.username or '',
                'reason': action.reason or '',
                'collection_owner': action.collection_owner or '',
                'make_owner': action.make_owner or '',
                'supplier': action.supplier or '',
                'collection': action.collection or '',
                'status_filter': action.status_filter or '',
                'po_number': po_num,
                'po_date': item.get('po_date') or item.get('order_date') or '',
                'order_weight': item.get('total_weight') or item.get('order_weight') or 0.0,
                'order_piece': item.get('order_piece') or 0,
                'schedules': schedules_str
            })

    # ── 3. Create Workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pending Acceptance Actions'

    # Styles
    PRIMARY_FILL = PatternFill('solid', fgColor='1E3A5F')
    HEADER_FILL = PatternFill('solid', fgColor='2563EB')
    ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
    
    TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    HEADER_FONT = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    DATA_FONT = Font(name='Calibri', size=9, color='1E293B')
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    COLUMNS = [
        ('Action ID',    10, 'center'),
        ('Action Type',  12, 'center'),
        ('Action Date',  18, 'center'),
        ('Username',     18, 'left'),
        ('Reason',       30, 'left'),
        ('Coll. Owner',  20, 'left'),
        ('Make Owner',   20, 'left'),
        ('Supplier',     25, 'left'),
        ('Collection',   20, 'left'),
        ('PO Number',    18, 'left'),
        ('PO Date',      15, 'center'),
        ('PO Weight',    14, 'right'),
        ('PO Pieces',    12, 'right'),
        ('Schedules (Fulfillment)', 40, 'left')
    ]

    TOTAL_COLS = len(COLUMNS)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1, value='Pending Acceptance Feedback Action Details Export')
    title_cell.font = TITLE_FONT
    title_cell.fill = PRIMARY_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    for col_idx, (label, width, align) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    # Data
    for row_idx, row_data in enumerate(po_rows, start=3):
        values = [
            row_data['action_id'],
            row_data['action_type'],
            row_data['action_date'],
            row_data['username'],
            row_data['reason'],
            row_data['collection_owner'],
            row_data['make_owner'],
            row_data['supplier'],
            row_data['collection'],
            row_data['po_number'],
            row_data['po_date'],
            row_data['order_weight'],
            row_data['order_piece'],
            row_data['schedules']
        ]
        alt = (row_idx % 2 == 0)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal=COLUMNS[col_idx-1][2], vertical='center')
            cell.border = THIN_BORDER
            if alt:
                cell.fill = ALT_FILL
            
            # Number formatting
            if col_idx == 12: # Weight
                cell.number_format = '#,##0.000'
            elif col_idx == 13: # Piece
                cell.number_format = '#,##0'

    # Save
    now_ist = datetime.now(IST)
    timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f'pending_acceptance_export_{timestamp}.xlsx'
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    logger.info(f'Export saved: {filepath} ({len(po_rows)} records)')
    return filename


def generate_provision_allocation_export(filters: dict) -> str:
    """
    Generate Provision Allocation Summary Excel export based on filters and search,
    rendering it in a beautiful, multi-column grid dashboard layout matching the UI.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(f"openpyxl not installed: {e}")

    from app.extensions import db
    from sqlalchemy import text

    _ensure_exports_dir()

    # 1. Parse and build query params
    location = filters.get('location', '')
    purity = filters.get('purity', '')
    classification = filters.get('classification', '')
    make = filters.get('make', '')
    collection = filters.get('collection', '')
    section = filters.get('section', '')
    prov_type = filters.get('prov_type', '')
    provision_mode = filters.get('provision_mode', '')
    branch_type = filters.get('branch_type', '')
    branch_status = filters.get('branch_status', '')
    business_head = filters.get('business_head', '')
    state = filters.get('state', '')
    search = filters.get('search', '').strip()

    params = {
        'location': location if location else None,
        'purity': float(purity) if purity else None,
        'classification': classification if classification else None,
        'make': make if make else None,
        'collection': collection if collection else None,
        'section': section if section else None,
        'prov_type': prov_type if prov_type else None,
        'provision_mode': provision_mode if provision_mode else None,
        'branch_type': branch_type if branch_type else None,
        'branch_status': branch_status if branch_status else None,
        'business_head': business_head if business_head else None,
        'state': state if state else None
    }

    # Use the exact same query as the partial endpoint
    query = """
WITH base AS (
    SELECT *
    FROM provision_stock_raw_snapshot
    WHERE 
        (:location IS NULL OR location = ANY(string_to_array(CAST(:location AS text), ',')))
        AND (:purity IS NULL OR purity = :purity)
        AND (:classification IS NULL OR classification = :classification)
        AND (:make IS NULL OR make = ANY(string_to_array(CAST(:make AS text), ',')))
        AND (:collection IS NULL OR collection = :collection)
        AND (:section IS NULL OR section = :section)
        AND (:prov_type IS NULL OR prov_type = :prov_type)
        AND (:provision_mode IS NULL OR provision_mode_filter = :provision_mode)
        AND (:branch_type IS NULL OR branch_type = ANY(string_to_array(CAST(:branch_type AS text), ',')))
        AND (:branch_status IS NULL OR branch_status = ANY(string_to_array(CAST(:branch_status AS text), ',')))
        AND (:business_head IS NULL OR business_head_name = :business_head)
        AND (:state IS NULL OR state = ANY(string_to_array(CAST(:state AS text), ',')))
),

global_total AS (
    SELECT
        COALESCE(SUM(prov_gr_wt), 0) AS total_prov_wt
    FROM base
),

location_summary AS (
    SELECT
        CASE 
            WHEN COUNT(DISTINCT location) > 4 THEN COUNT(DISTINCT location)::text || '+ Location'
            ELSE (SELECT STRING_AGG(loc, ', ') FROM (SELECT DISTINCT location AS loc FROM base ORDER BY loc) s)
        END::text AS location,
        'Location Summary'::text AS report_section,
        CASE 
            WHEN COUNT(DISTINCT location) > 4 THEN COUNT(DISTINCT location)::text || '+ Location'
            ELSE (SELECT STRING_AGG(loc, ', ') FROM (SELECT DISTINCT location AS loc FROM base ORDER BY loc) s)
        END::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        SUM(prov_pieces) AS pcs,
        SUM(prov_gr_wt) AS gr_wt,
        100.00::numeric AS percent,
        1 AS section_sort,
        1 AS row_sort
    FROM base
),

purity_wise AS (
    SELECT
        'ALL'::text AS location,
        'Purity Wise'::text AS report_section,
        COALESCE(b.purity::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        NULL::numeric AS pcs,
        SUM(b.prov_gr_wt) AS gr_wt,
        ROUND(
            CASE
                WHEN gt.total_prov_wt = 0 THEN 0
                ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
            END,
            2
        ) AS percent,
        2 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.purity) AS row_sort
    FROM base b
    CROSS JOIN global_total gt
    GROUP BY b.purity, gt.total_prov_wt
),

classification_wise AS (
    SELECT
        'ALL'::text AS location,
        x.report_section,
        x.report_label,
        x.classification,
        x.sub_classification,
        x.is_parent,
        x.pcs,
        x.gr_wt,
        x.percent,
        3 AS section_sort,
        ROW_NUMBER() OVER (
            ORDER BY
                x.classification,
                x.level_order,
                x.sub_classification NULLS FIRST
        ) AS row_sort
    FROM (
        SELECT
            'Classification Wise'::text AS report_section,
            COALESCE(b.classification::text, 'Unknown') AS report_label,
            b.classification::text AS classification,
            NULL::text AS sub_classification,
            1 AS is_parent,
            NULL::numeric AS pcs,
            SUM(b.prov_gr_wt) AS gr_wt,
            ROUND(
                CASE
                    WHEN gt.total_prov_wt = 0 THEN 0
                    ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
                END,
                2
            ) AS percent,
            0 AS level_order
        FROM base b
        CROSS JOIN global_total gt
        GROUP BY b.classification, gt.total_prov_wt

        UNION ALL

        SELECT
            'Classification Wise'::text AS report_section,
            '   ' || COALESCE(b.sub_classification::text, 'Unknown') AS report_label,
            b.classification::text AS classification,
            COALESCE(b.sub_classification::text, 'Unknown') AS sub_classification,
            0 AS is_parent,
            NULL::numeric AS pcs,
            SUM(b.prov_gr_wt) AS gr_wt,
            ROUND(
                CASE
                    WHEN gt.total_prov_wt = 0 THEN 0
                    ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
                END,
                2
            ) AS percent,
            1 AS level_order
        FROM base b
        CROSS JOIN global_total gt
        GROUP BY b.classification, b.sub_classification, gt.total_prov_wt
    ) x
),

make_wise AS (
    SELECT
        'ALL'::text AS location,
        'Make Wise'::text AS report_section,
        COALESCE(b.make::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        NULL::numeric AS pcs,
        SUM(b.prov_gr_wt) AS gr_wt,
        ROUND(
            CASE
                WHEN gt.total_prov_wt = 0 THEN 0
                ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
            END,
            2
        ) AS percent,
        4 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.make) AS row_sort
    FROM base b
    CROSS JOIN global_total gt
    GROUP BY b.make, gt.total_prov_wt
),

prov_type_wise AS (
    SELECT
        'ALL'::text AS location,
        'Provision Type Wise'::text AS report_section,
        COALESCE(b.prov_type::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        NULL::numeric AS pcs,
        SUM(b.prov_gr_wt) AS gr_wt,
        ROUND(
            CASE
                WHEN gt.total_prov_wt = 0 THEN 0
                ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
            END,
            2
        ) AS percent,
        5 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.prov_type) AS row_sort
    FROM base b
    CROSS JOIN global_total gt
    GROUP BY b.prov_type, gt.total_prov_wt
),

section_wise AS (
    SELECT
        'ALL'::text AS location,
        'Section Wise'::text AS report_section,
        COALESCE(b.section::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        SUM(b.prov_pieces) AS pcs,
        SUM(b.prov_gr_wt) AS gr_wt,
        ROUND(
            CASE
                WHEN gt.total_prov_wt = 0 THEN 0
                ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
            END,
            2
        ) AS percent,
        6 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.section) AS row_sort
    FROM base b
    CROSS JOIN global_total gt
    GROUP BY b.section, gt.total_prov_wt
),

provision_mode_wise AS (
    SELECT
        'ALL'::text AS location,
        'Provision Mode Wise'::text AS report_section,
        COALESCE(b.provision_mode_filter::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        NULL::numeric AS pcs,
        SUM(b.prov_gr_wt) AS gr_wt,
        ROUND(
            CASE
                WHEN gt.total_prov_wt = 0 THEN 0
                ELSE SUM(b.prov_gr_wt) * 100.0 / gt.total_prov_wt
            END,
            2
        ) AS percent,
        7 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.provision_mode_filter) AS row_sort
    FROM base b
    CROSS JOIN global_total gt
    GROUP BY b.provision_mode_filter, gt.total_prov_wt
),

provision_mode_count AS (
    SELECT
        'ALL'::text AS location,
        'Provision Mode Count'::text AS report_section,
        COALESCE(b.provision_mode_filter::text, 'Unknown') AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        1 AS is_parent,
        SUM(b.prov_pieces) AS pcs,
        NULL::numeric AS gr_wt,
        NULL::numeric AS percent,
        8 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY b.provision_mode_filter) AS row_sort
    FROM base b
    GROUP BY b.provision_mode_filter
),

combined_report AS (
    SELECT * FROM location_summary
    UNION ALL
    SELECT * FROM purity_wise
    UNION ALL
    SELECT * FROM classification_wise
    UNION ALL
    SELECT * FROM make_wise
    UNION ALL
    SELECT * FROM prov_type_wise
    UNION ALL
    SELECT * FROM section_wise
    UNION ALL
    SELECT * FROM provision_mode_wise
    UNION ALL
    SELECT * FROM provision_mode_count
)

SELECT
    location,
    report_section,
    report_label,
    classification,
    sub_classification,
    is_parent,
    pcs,
    gr_wt AS grossweight,
    percent,
    section_sort,
    row_sort
FROM combined_report
ORDER BY
    location,
    section_sort,
    row_sort
    """

    result = db.session.execute(text(query), params)
    all_rows = [dict(r._mapping) for r in result]

    # Search filter logic
    if search:
        s_lower = search.lower()
        all_rows = [
            r for r in all_rows
            if (s_lower in (r.get('report_label') or '').lower() or
                s_lower in (r.get('report_section') or '').lower())
        ]

    # Group into segments
    segments = {}
    for row in all_rows:
        if row.get('report_label') == 'Grand Total':
            continue
        section = row.get('report_section')
        if not section:
            continue
        if section not in segments:
            segments[section] = []
        segments[section].append(row)

    # Sort each segment by percent descending, EXCEPT Classification Wise
    for section in list(segments.keys()):
        if section != 'Classification Wise':
            segments[section] = sorted(
                segments[section],
                key=lambda x: x.get('percent') or 0.0,
                reverse=True
            )

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Provision Allocation Summary'
    ws.views.sheetView[0].showGridLines = True

    # Styling colors
    PRIMARY_FILL = PatternFill('solid', fgColor='1E3A5F') # Navy
    HEADER_FILL = PatternFill('solid', fgColor='FCE4EC') # Light pink as in UI
    TOTAL_FILL = PatternFill('solid', fgColor='D6D6D6') # Gray
    CAT_FILL = PatternFill('solid', fgColor='E3F2FD') # Light blue
    ALT_FILL = PatternFill('solid', fgColor='F8FAFC')

    TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    SECTION_HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='1E3A5F')
    HEADER_FONT = Font(name='Calibri', bold=True, size=9, color='374151')
    DATA_FONT = Font(name='Calibri', size=9, color='1F2937')
    BOLD_FONT = Font(name='Calibri', bold=True, size=9, color='111827')

    THIN_BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Column setups:
    # Col 1: A, B, C
    # Spacer: D
    # Col 2: E, F, G
    # Spacer: H
    # Col 3: I, J, K
    # Spacer: L
    # Col 4: M, N, O, P
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 4
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 16
    ws.column_dimensions['P'].width = 12

    # Title Block
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = 'Provision Allocation Summary Report'
    title_cell.font = TITLE_FONT
    title_cell.fill = PRIMARY_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 4 tracking cursors for columns
    col_1_row = 4
    col_2_row = 4
    col_3_row = 4
    col_4_row = 4

    # ── Grid Column 1 ────────────────────────────────────────────────────────

    # 1. Location Summary
    if 'Location Summary' in segments and segments['Location Summary']:
        loc_row = segments['Location Summary'][0]
        # Draw Location Summary table
        ws.merge_cells(start_row=col_1_row, start_column=1, end_row=col_1_row, end_column=3)
        cell = ws.cell(row=col_1_row, column=1, value=f"Location Summary: {loc_row.get('location') or '-'}")
        cell.font = BOLD_FONT
        cell.fill = CAT_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        # Need border on merged cells too
        for col in range(1, 4):
            ws.cell(row=col_1_row, column=col).border = THIN_BORDER
        col_1_row += 2

    # Helper function to render a standard (Label, Gr.Wt, %) table
    def render_excel_std_table(start_row, start_col, title, rows, category_labels=None):
        # 1. Title Header row
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = SECTION_HEADER_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        start_row += 1

        # 2. Table Column Headers
        headers = [title, 'Gr.Wt', '%']
        for offset, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col+offset, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='left' if offset==0 else 'right', vertical='center')
            cell.border = THIN_BORDER
        start_row += 1

        # 3. Data rows
        totals_wt = 0.0
        totals_pct = 0.0
        has_grand = False
        use_cats = False

        if category_labels:
            for r in rows:
                if r.get('report_label') in category_labels:
                    use_cats = True
                    break

        curr_row = start_row
        for idx, r in enumerate(rows):
            label = r.get('report_label') or ''
            wt = _safe_float(r.get('grossweight'))
            pct = _safe_float(r.get('percent'))
            is_parent = r.get('is_parent') if r.get('is_parent') is not None else 1

            is_gt = label == 'Grand Total'
            is_cat = category_labels is not None and label in category_labels

            if is_gt:
                has_grand = True
                # Print Grand Total Row
                cell_lbl = ws.cell(row=curr_row, column=start_col, value=label)
                cell_wt = ws.cell(row=curr_row, column=start_col+1, value=wt)
                cell_pct = ws.cell(row=curr_row, column=start_col+2, value=pct / 100.0)

                for c in range(start_col, start_col+3):
                    ws.cell(row=curr_row, column=c).font = BOLD_FONT
                    ws.cell(row=curr_row, column=c).fill = TOTAL_FILL
                    ws.cell(row=curr_row, column=c).border = THIN_BORDER

                cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
                cell_wt.alignment = Alignment(horizontal='right', vertical='center')
                cell_pct.alignment = Alignment(horizontal='right', vertical='center')

                cell_wt.number_format = '#,##0.000'
                cell_pct.number_format = '0%'
            else:
                # Add to totals
                if use_cats:
                    if is_cat:
                        totals_wt += wt
                        totals_pct += pct
                else:
                    if is_parent != 0:
                        totals_wt += wt
                        totals_pct += pct

                cell_lbl = ws.cell(row=curr_row, column=start_col, value=label)
                cell_wt = ws.cell(row=curr_row, column=start_col+1, value=wt)
                cell_pct = ws.cell(row=curr_row, column=start_col+2, value=pct / 100.0)

                alt = (idx % 2 == 1)
                row_fill = ALT_FILL if alt else PatternFill(fill_type=None)
                if is_cat:
                    row_fill = CAT_FILL

                for c in range(start_col, start_col+3):
                    cell_curr = ws.cell(row=curr_row, column=c)
                    cell_curr.border = THIN_BORDER
                    if row_fill.fill_type:
                        cell_curr.fill = row_fill
                    if is_cat:
                        cell_curr.font = BOLD_FONT
                    else:
                        cell_curr.font = DATA_FONT

                # Indentation for classification child rows
                if is_parent == 0:
                    cell_lbl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                else:
                    cell_lbl.alignment = Alignment(horizontal='left', vertical='center')

                cell_wt.alignment = Alignment(horizontal='right', vertical='center')
                cell_pct.alignment = Alignment(horizontal='right', vertical='center')

                cell_wt.number_format = '#,##0.000'
                cell_pct.number_format = '0%'

            curr_row += 1

        # Append Grand Total if missing
        if not has_grand and len(rows) > 0:
            cell_lbl = ws.cell(row=curr_row, column=start_col, value='Grand Total')
            cell_wt = ws.cell(row=curr_row, column=start_col+1, value=totals_wt)
            cell_pct = ws.cell(row=curr_row, column=start_col+2, value=1.0) # 100%

            for c in range(start_col, start_col+3):
                cell_curr = ws.cell(row=curr_row, column=c)
                cell_curr.font = BOLD_FONT
                cell_curr.fill = TOTAL_FILL
                cell_curr.border = THIN_BORDER

            cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
            cell_wt.alignment = Alignment(horizontal='right', vertical='center')
            cell_pct.alignment = Alignment(horizontal='right', vertical='center')

            cell_wt.number_format = '#,##0.000'
            cell_pct.number_format = '0%'
            curr_row += 1

        return curr_row

    # 2. Purity Wise
    if 'Purity Wise' in segments:
        col_1_row = render_excel_std_table(col_1_row, 1, 'Purity Wise', segments['Purity Wise'])
        col_1_row += 2

    # 3. Classification Wise
    if 'Classification Wise' in segments:
        col_1_row = render_excel_std_table(col_1_row, 1, 'Classification Wise', segments['Classification Wise'], category_labels=['BRAND', 'GENERIC', 'LIFE STYLE'])

    # ── Grid Column 2 ────────────────────────────────────────────────────────

    # 4. Make Wise
    if 'Make Wise' in segments:
        col_2_row = render_excel_std_table(col_2_row, 5, 'Make Wise', segments['Make Wise'])

    # ── Grid Column 3 ────────────────────────────────────────────────────────

    # 5. Provision Type Wise
    if 'Provision Type Wise' in segments:
        col_3_row = render_excel_std_table(col_3_row, 9, 'Provision Type Wise', segments['Provision Type Wise'], category_labels=['GeneralProvision', 'ManagerProvision', 'SetProvision'])
        col_3_row += 2

    # 6. Provision Mode Wise
    if 'Provision Mode Wise' in segments:
        col_3_row = render_excel_std_table(col_3_row, 9, 'Provision Mode Wise', segments['Provision Mode Wise'], category_labels=['MatchingSet', 'Set'])

    # ── Grid Column 4 ────────────────────────────────────────────────────────

    # 7. Section Wise
    if 'Section Wise' in segments:
        sec_start_row = col_4_row
        ws.merge_cells(start_row=sec_start_row, start_column=13, end_row=sec_start_row, end_column=16)
        title_cell = ws.cell(row=sec_start_row, column=13, value='Section Wise')
        title_cell.font = SECTION_HEADER_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        sec_start_row += 1

        headers = ['Section', 'Pcs', 'Gr.Wt', '%']
        for offset, h in enumerate(headers):
            cell = ws.cell(row=sec_start_row, column=13+offset, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='left' if offset==0 else 'right', vertical='center')
            cell.border = THIN_BORDER
        sec_start_row += 1

        totals_pcs = 0
        totals_wt = 0.0
        has_grand = False

        for idx, r in enumerate(segments['Section Wise']):
            label = r.get('report_label') or ''
            pcs = _safe_int(r.get('pcs'))
            wt = _safe_float(r.get('grossweight'))
            pct = _safe_float(r.get('percent'))

            is_gt = label == 'Grand Total'
            if is_gt:
                has_grand = True
                cell_lbl = ws.cell(row=sec_start_row, column=13, value=label)
                cell_pcs = ws.cell(row=sec_start_row, column=14, value=pcs)
                cell_wt = ws.cell(row=sec_start_row, column=15, value=wt)
                cell_pct = ws.cell(row=sec_start_row, column=16, value=pct / 100.0)

                for c in range(13, 17):
                    ws.cell(row=sec_start_row, column=c).font = BOLD_FONT
                    ws.cell(row=sec_start_row, column=c).fill = TOTAL_FILL
                    ws.cell(row=sec_start_row, column=c).border = THIN_BORDER

                cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
                cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
                cell_wt.alignment = Alignment(horizontal='right', vertical='center')
                cell_pct.alignment = Alignment(horizontal='right', vertical='center')

                cell_pcs.number_format = '#,##0'
                cell_wt.number_format = '#,##0.000'
                cell_pct.number_format = '0%'
            else:
                totals_pcs += pcs
                totals_wt += wt

                cell_lbl = ws.cell(row=sec_start_row, column=13, value=label)
                cell_pcs = ws.cell(row=sec_start_row, column=14, value=pcs)
                cell_wt = ws.cell(row=sec_start_row, column=15, value=wt)
                cell_pct = ws.cell(row=sec_start_row, column=16, value=pct / 100.0)

                alt = (idx % 2 == 1)
                row_fill = ALT_FILL if alt else PatternFill(fill_type=None)

                for c in range(13, 17):
                    cell_curr = ws.cell(row=sec_start_row, column=c)
                    cell_curr.font = DATA_FONT
                    cell_curr.border = THIN_BORDER
                    if row_fill.fill_type:
                        cell_curr.fill = row_fill

                cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
                cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
                cell_wt.alignment = Alignment(horizontal='right', vertical='center')
                cell_pct.alignment = Alignment(horizontal='right', vertical='center')

                cell_pcs.number_format = '#,##0'
                cell_wt.number_format = '#,##0.000'
                cell_pct.number_format = '0%'

            sec_start_row += 1

        if not has_grand and len(segments['Section Wise']) > 0:
            cell_lbl = ws.cell(row=sec_start_row, column=13, value='Grand Total')
            cell_pcs = ws.cell(row=sec_start_row, column=14, value=totals_pcs)
            cell_wt = ws.cell(row=sec_start_row, column=15, value=totals_wt)
            cell_pct = ws.cell(row=sec_start_row, column=16, value=1.0)

            for c in range(13, 17):
                cell_curr = ws.cell(row=sec_start_row, column=c)
                cell_curr.font = BOLD_FONT
                cell_curr.fill = TOTAL_FILL
                cell_curr.border = THIN_BORDER

            cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
            cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
            cell_wt.alignment = Alignment(horizontal='right', vertical='center')
            cell_pct.alignment = Alignment(horizontal='right', vertical='center')

            cell_pcs.number_format = '#,##0'
            cell_wt.number_format = '#,##0.000'
            cell_pct.number_format = '0%'
            sec_start_row += 1

        col_4_row = sec_start_row + 2

    # 8. Provision Mode Count
    if 'Provision Mode Count' in segments:
        cnt_start_row = col_4_row
        ws.merge_cells(start_row=cnt_start_row, start_column=13, end_row=cnt_start_row, end_column=14)
        title_cell = ws.cell(row=cnt_start_row, column=13, value='Provision Mode Count')
        title_cell.font = SECTION_HEADER_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        cnt_start_row += 1

        headers = ['Provision Mode', 'Count']
        for offset, h in enumerate(headers):
            cell = ws.cell(row=cnt_start_row, column=13+offset, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='left' if offset==0 else 'right', vertical='center')
            cell.border = THIN_BORDER
        cnt_start_row += 1

        totals_pcs = 0
        has_grand = False
        category_labels = ['BRIDAL SET', 'COMBO SET', 'GeneralProvision', 'ManagerProvision', 'SetProvision']
        use_cats = False

        for r in segments['Provision Mode Count']:
            if r.get('report_label') in category_labels:
                use_cats = True
                break

        for idx, r in enumerate(segments['Provision Mode Count']):
            label = r.get('report_label') or ''
            pcs = _safe_int(r.get('pcs'))

            is_gt = label == 'Grand Total'
            is_cat = label in category_labels

            if is_gt:
                has_grand = True
                cell_lbl = ws.cell(row=cnt_start_row, column=13, value=label)
                cell_pcs = ws.cell(row=cnt_start_row, column=14, value=pcs)

                for c in range(13, 15):
                    ws.cell(row=cnt_start_row, column=c).font = BOLD_FONT
                    ws.cell(row=cnt_start_row, column=c).fill = TOTAL_FILL
                    ws.cell(row=cnt_start_row, column=c).border = THIN_BORDER

                cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
                cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
                cell_pcs.number_format = '#,##0'
            else:
                if use_cats:
                    if is_cat:
                        totals_pcs += pcs
                else:
                    totals_pcs += pcs

                cell_lbl = ws.cell(row=cnt_start_row, column=13, value=label)
                cell_pcs = ws.cell(row=cnt_start_row, column=14, value=pcs)

                alt = (idx % 2 == 1)
                row_fill = ALT_FILL if alt else PatternFill(fill_type=None)
                if is_cat:
                    row_fill = CAT_FILL

                for c in range(13, 15):
                    cell_curr = ws.cell(row=cnt_start_row, column=c)
                    cell_curr.border = THIN_BORDER
                    if row_fill.fill_type:
                        cell_curr.fill = row_fill
                    if is_cat:
                        cell_curr.font = BOLD_FONT
                    else:
                        cell_curr.font = DATA_FONT

                if not (is_cat or is_gt) and use_cats:
                    cell_lbl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                else:
                    cell_lbl.alignment = Alignment(horizontal='left', vertical='center')

                cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
                cell_pcs.number_format = '#,##0'

            cnt_start_row += 1

        if not has_grand and len(segments['Provision Mode Count']) > 0:
            cell_lbl = ws.cell(row=cnt_start_row, column=13, value='Grand Total')
            cell_pcs = ws.cell(row=cnt_start_row, column=14, value=totals_pcs)

            for c in range(13, 15):
                cell_curr = ws.cell(row=cnt_start_row, column=c)
                cell_curr.font = BOLD_FONT
                cell_curr.fill = TOTAL_FILL
                cell_curr.border = THIN_BORDER

            cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
            cell_pcs.alignment = Alignment(horizontal='right', vertical='center')
            cell_pcs.number_format = '#,##0'
            cnt_start_row += 1

        col_4_row = cnt_start_row

    # 9. Save
    now_ist = datetime.now(IST)
    timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f'provision_allocation_summary_{timestamp}.xlsx'
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    logger.info(f'Provision Allocation Summary export saved: {filepath}')
    return filename


def generate_location_physical_stock_status_export(filters: dict) -> str:
    """
    Generate Location Physical Stock Status Excel export based on filters and search,
    rendering it in a beautiful, multi-column grid dashboard layout matching the UI.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("openpyxl not installed: " + str(e))

    from app.extensions import db
    from sqlalchemy import text

    _ensure_exports_dir()

    # 1. Parse and build query params
    location = filters.get('location', '')
    state = filters.get('state', '')
    purity = filters.get('purity', '')
    classification = filters.get('classification', '')
    make = filters.get('make', '')
    collection = filters.get('collection', '')
    section = filters.get('section', '')
    prov_type = filters.get('prov_type', '')
    provision_mode = filters.get('provision_mode', '')
    branch_type = filters.get('branch_type', '')
    branch_status = filters.get('branch_status', '')
    business_head = filters.get('business_head', '')
    bh_emp_code = filters.get('bh_emp_code')
    authorized_branch_ids = filters.get('authorized_branch_ids')
    search = filters.get('search', '').strip()
    sort_by = filters.get('sort_by', '')
    sort_order = filters.get('sort_order', 'asc')

    params = {
        'location': location if location else None,
        'state': state if state else None,
        'purity': purity if purity else None,
        'classification': classification if classification else None,
        'make': make if make else None,
        'collection': collection if collection else None,
        'section': section if section else None,
        'prov_type': prov_type if prov_type else None,
        'provision_mode': provision_mode if provision_mode else None,
        'branch_type': branch_type if branch_type else None,
        'branch_status': branch_status if branch_status else None,
        'business_head': business_head if business_head else None,
        'bh_emp_code': bh_emp_code,
        'authorized_branch_ids': authorized_branch_ids
    }

    # Query matching routes/location_physical_stock_status.py
    query = """
WITH base AS (
    SELECT *
    FROM provision_stock_raw_snapshot
    WHERE 
        (:location IS NULL OR location = ANY(string_to_array(CAST(:location AS text), ',')))
        AND (:state IS NULL OR state = ANY(string_to_array(CAST(:state AS text), ',')))
        AND (:purity IS NULL OR purity = ANY(string_to_array(CAST(:purity AS text), ',')::numeric[]))
        AND (:classification IS NULL OR classification = ANY(string_to_array(CAST(:classification AS text), ',')))
        AND (:make IS NULL OR make = ANY(string_to_array(CAST(:make AS text), ',')))
        AND (:collection IS NULL OR collection = ANY(string_to_array(CAST(:collection AS text), ',')))
        AND (:section IS NULL OR section = ANY(string_to_array(CAST(:section AS text), ',')))
        AND (:prov_type IS NULL OR prov_type = ANY(string_to_array(CAST(:prov_type AS text), ',')))
        AND (:provision_mode IS NULL OR provision_mode_filter = ANY(string_to_array(CAST(:provision_mode AS text), ',')))
        AND (:branch_type IS NULL OR branch_type = ANY(string_to_array(CAST(:branch_type AS text), ',')))
        AND (:branch_status IS NULL OR branch_status = ANY(string_to_array(CAST(:branch_status AS text), ',')))
        AND (:business_head IS NULL OR business_head_name = ANY(string_to_array(CAST(:business_head AS text), ',')))
        AND (:bh_emp_code IS NULL OR business_head_emp_code = :bh_emp_code)
        AND (:authorized_branch_ids IS NULL OR branch_id = ANY(string_to_array(CAST(:authorized_branch_ids AS text), ',')::integer[]))
),
location_summary AS (
    SELECT
        location::text AS location,
        'Location Summary'::text AS report_section,
        location::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        SUM(prov_pieces) AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        1 AS section_sort,
        1 AS row_sort
    FROM base
    GROUP BY location
),
purity_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        'Purity Wise'::text AS report_section,
        purity::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        NULL::numeric AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        2 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY purity) AS row_sort
    FROM base
    GROUP BY purity
),
classification_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        x.report_section,
        x.report_label,
        x.classification,
        x.sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        x.is_parent,
        x.prov_pcs,
        x.prov_gr_wt,
        x.in_shop_wt,
        x.ordered_wt,
        x.in_transit_wt,
        x.short_excess_wt,
        x.percent,
        3 AS section_sort,
        ROW_NUMBER() OVER (
            ORDER BY x.classification, x.level_order, x.sub_classification NULLS FIRST
        ) AS row_sort
    FROM (
        SELECT
            'Classification Wise'::text AS report_section,
            classification::text AS report_label,
            classification::text AS classification,
            NULL::text AS sub_classification,
            1 AS is_parent,
            NULL::numeric AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            0 AS level_order
        FROM base
        GROUP BY classification
 
        UNION ALL
 
        SELECT
            'Classification Wise'::text AS report_section,
            '   ' || COALESCE(sub_classification::text, 'Unknown') AS report_label,
            classification::text AS classification,
            COALESCE(sub_classification::text, 'Unknown') AS sub_classification,
            0 AS is_parent,
            NULL::numeric AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            1 AS level_order
        FROM base
        GROUP BY classification, sub_classification
    ) x
),
collection_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        x.report_section,
        x.report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        x.collection,
        x.sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        x.is_parent,
        x.prov_pcs,
        x.prov_gr_wt,
        x.in_shop_wt,
        x.ordered_wt,
        x.in_transit_wt,
        x.short_excess_wt,
        x.percent,
        10 AS section_sort,
        ROW_NUMBER() OVER (
            ORDER BY x.collection, x.level_order, x.sub_section NULLS FIRST
        ) AS row_sort
    FROM (
        SELECT
            'Collection Wise'::text AS report_section,
            collection::text AS report_label,
            collection::text AS collection,
            NULL::text AS sub_section,
            1 AS is_parent,
            NULL::numeric AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            0 AS level_order
        FROM base
        GROUP BY collection
 
        UNION ALL
 
        SELECT
            'Collection Wise'::text AS report_section,
            '   ' || COALESCE(sub_section::text, 'Unknown') AS report_label,
            collection::text AS collection,
            COALESCE(sub_section::text, 'Unknown') AS sub_section,
            0 AS is_parent,
            NULL::numeric AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            1 AS level_order
        FROM base
        GROUP BY collection, sub_section
    ) x
),
section_details_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        x.report_section,
        x.report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        x.sec_name,
        x.typ_name,
        x.is_parent,
        x.prov_pcs,
        x.prov_gr_wt,
        x.in_shop_wt,
        x.ordered_wt,
        x.in_transit_wt,
        x.short_excess_wt,
        x.percent,
        9 AS section_sort,
        ROW_NUMBER() OVER (
            ORDER BY x.sec_name, x.level_order, x.typ_name NULLS FIRST
        ) AS row_sort
    FROM (
        SELECT
            'Section Details'::text AS report_section,
            section::text AS report_label,
            section::text AS sec_name,
            NULL::text AS typ_name,
            1 AS is_parent,
            SUM(prov_pieces) AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            0 AS level_order
        FROM base
        GROUP BY section
 
        UNION ALL
 
        SELECT
            'Section Details'::text AS report_section,
            '   ' || COALESCE(type::text, 'Unknown') AS report_label,
            section::text AS sec_name,
            COALESCE(type::text, 'Unknown') AS typ_name,
            0 AS is_parent,
            SUM(prov_pieces) AS prov_pcs,
            SUM(prov_gr_wt) AS prov_gr_wt,
            SUM(in_shop_wt) AS in_shop_wt,
            SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
            SUM(in_transit_wt) AS in_transit_wt,
            SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
            ROUND(
                CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
            ) AS percent,
            1 AS level_order
        FROM base
        GROUP BY section, type
    ) x
),
make_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        'Make Wise'::text AS report_section,
        make::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        NULL::numeric AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        5 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY make) AS row_sort
    FROM base
    GROUP BY make
),
prov_type_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        'Provision Type Wise'::text AS report_section,
        prov_type::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        NULL::numeric AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        6 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY prov_type) AS row_sort
    FROM base
    GROUP BY prov_type
),
section_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        'Section Wise'::text AS report_section,
        section::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        SUM(prov_pieces) AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        7 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY section) AS row_sort
    FROM base
    GROUP BY section
),
provision_mode_wise AS (
    SELECT
        'SUMMARY'::text AS location,
        'Provision Mode Wise'::text AS report_section,
        provision_mode_filter::text AS report_label,
        NULL::text AS classification,
        NULL::text AS sub_classification,
        NULL::text AS collection,
        NULL::text AS sub_section,
        NULL::text AS sec_name,
        NULL::text AS typ_name,
        1 AS is_parent,
        NULL::numeric AS prov_pcs,
        SUM(prov_gr_wt) AS prov_gr_wt,
        SUM(in_shop_wt) AS in_shop_wt,
        SUM(COALESCE(order_only_wt, 0) + COALESCE(req_only, 0)) AS ordered_wt,
        SUM(in_transit_wt) AS in_transit_wt,
        SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt) AS short_excess_wt,
        ROUND(
            CASE WHEN SUM(prov_gr_wt) = 0 THEN 0 ELSE (SUM(in_shop_wt) + SUM(in_transit_wt) - SUM(prov_gr_wt)) * 100.0 / SUM(prov_gr_wt) END, 2
        ) AS percent,
        8 AS section_sort,
        ROW_NUMBER() OVER (ORDER BY provision_mode_filter) AS row_sort
    FROM base
    GROUP BY provision_mode_filter
),
combined_report AS (
    SELECT * FROM location_summary
    UNION ALL
    SELECT * FROM purity_wise
    UNION ALL
    SELECT * FROM classification_wise
    UNION ALL
    SELECT * FROM collection_wise
    UNION ALL
    SELECT * FROM make_wise
    UNION ALL
    SELECT * FROM prov_type_wise
    UNION ALL
    SELECT * FROM section_wise
    UNION ALL
    SELECT * FROM provision_mode_wise
    UNION ALL
    SELECT * FROM section_details_wise
)
SELECT
    location,
    report_section,
    report_label,
    classification,
    sub_classification,
    collection,
    sub_section,
    sec_name,
    typ_name,
    is_parent,
    prov_pcs,
    prov_gr_wt,
    in_shop_wt,
    ordered_wt,
    in_transit_wt,
    short_excess_wt,
    percent,
    section_sort,
    row_sort
FROM combined_report
ORDER BY
    location,
    section_sort,
    row_sort
    """

    result = db.session.execute(text(query), params)
    all_rows = [dict(r._mapping) for r in result]

    # In-memory search filter
    if search:
        s_lower = search.lower()
        all_rows = [
            r for r in all_rows
            if (s_lower in (r.get('report_label') or '').lower() or
                s_lower in (r.get('report_section') or '').lower())
        ]

    # In-memory sorting (copy exactly from partial route)
    numeric_cols = ['prov_pcs', 'prov_gr_wt', 'in_shop_wt', 'ordered_wt', 'in_transit_wt', 'short_excess_wt', 'percent']
    if sort_by in numeric_cols:
        sections = {}
        for row in all_rows:
            section_name = row['report_section']
            if section_name not in sections:
                sections[section_name] = []
            sections[section_name].append(row)
        
        all_sorted_rows = []
        sorted_section_names = sorted(sections.keys(), key=lambda s: sections[s][0]['section_sort'])
        
        for s in sorted_section_names:
            sec_rows = sections[s]
            if s in ['Classification Wise', 'Collection Wise', 'Section Details']:
                key = 'classification' if s == 'Classification Wise' else ('collection' if s == 'Collection Wise' else 'sec_name')
                parents = [r for r in sec_rows if r['is_parent'] == 1]
                parents.sort(key=lambda r: float(r.get(sort_by) or 0), reverse=(sort_order == 'desc'))
                
                for p in parents:
                    all_sorted_rows.append(p)
                    children = [r for r in sec_rows if r['is_parent'] == 0 and r.get(key) == p.get(key)]
                    children.sort(key=lambda r: float(r.get(sort_by) or 0), reverse=(sort_order == 'desc'))
                    all_sorted_rows.extend(children)
            else:
                sec_rows.sort(key=lambda r: float(r.get(sort_by) or 0), reverse=(sort_order == 'desc'))
                all_sorted_rows.extend(sec_rows)
        
        all_rows = all_sorted_rows

    # Group into segments
    segments = {}
    for row in all_rows:
        if row.get('report_label') == 'Grand Total':
            continue
        section_name = row.get('report_section')
        if not section_name:
            continue
        if section_name not in segments:
            segments[section_name] = []
        segments[section_name].append(row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Location Stock Status'
    ws.views.sheetView[0].showGridLines = True

    # Styling colors
    PRIMARY_FILL = PatternFill('solid', fgColor='1E3A5F') # Navy
    HEADER_FILL = PatternFill('solid', fgColor='FCE4EC') # Light pink as in UI
    TOTAL_FILL = PatternFill('solid', fgColor='D6D6D6') # Gray
    CAT_FILL = PatternFill('solid', fgColor='E3F2FD') # Light blue
    ALT_FILL = PatternFill('solid', fgColor='F8FAFC')

    TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    SECTION_HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='1E3A5F')
    HEADER_FONT = Font(name='Calibri', bold=True, size=9, color='374151')
    DATA_FONT = Font(name='Calibri', size=9, color='1F2937')
    BOLD_FONT = Font(name='Calibri', bold=True, size=9, color='111827')

    THIN_BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Column setups
    # Column 1: A to G
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    
    # Spacer
    ws.column_dimensions['H'].width = 4
    
    # Column 2: I to P
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 14
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 14
    
    # Spacer
    ws.column_dimensions['Q'].width = 4
    
    # Column 3: R to X
    ws.column_dimensions['R'].width = 25
    ws.column_dimensions['S'].width = 14
    ws.column_dimensions['T'].width = 14
    ws.column_dimensions['U'].width = 14
    ws.column_dimensions['V'].width = 14
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 14

    # Title Block
    ws.merge_cells('A1:X1')
    title_cell = ws['A1']
    title_cell.value = 'Location Physical Stock Status Report'
    title_cell.font = TITLE_FONT
    title_cell.fill = PRIMARY_FILL
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 3 vertical tracking cursors for columns
    col_1_row = 4
    col_2_row = 4
    col_3_row = 4

    # Helper function to render a table in Excel
    def render_excel_table(start_row, start_col, title, rows, show_pcs=False, category_labels=None):
        col_span = 7 if show_pcs else 6
        # Title row
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+col_span)
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = SECTION_HEADER_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Border & Fill on title merged cells
        for c in range(start_col, start_col + col_span + 1):
            ws.cell(row=start_row, column=c).border = THIN_BORDER
            ws.cell(row=start_row, column=c).fill = CAT_FILL
        start_row += 1

        # Table Column Headers
        if show_pcs:
            headers = [title, 'Prov Pcs', 'Prov Gr.Wt', 'In Shop Wt', 'Transit Wt', 'Short/Excess', '%', 'Ordered Wt']
        else:
            headers = [title, 'Prov Gr.Wt', 'In Shop Wt', 'Transit Wt', 'Short/Excess', '%', 'Ordered Wt']

        for offset, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col+offset, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='left' if offset==0 else 'right', vertical='center')
            cell.border = THIN_BORDER
        start_row += 1

        # Data rows
        totals = {
            'pcs': 0,
            'prov_gr_wt': 0.0,
            'in_shop_wt': 0.0,
            'in_transit_wt': 0.0,
            'short_excess_wt': 0.0,
            'ordered_wt': 0.0
        }
        has_grand = False
        use_cats = False

        if category_labels:
            for r in rows:
                if r.get('report_label') in category_labels:
                    use_cats = True
                    break

        curr_row = start_row
        for idx, r in enumerate(rows):
            label = r.get('report_label') or ''
            is_parent = r.get('is_parent') if r.get('is_parent') is not None else 1

            is_gt = label == 'Grand Total'
            is_cat = category_labels is not None and label in category_labels

            pcs = _safe_int(r.get('prov_pcs')) if show_pcs else None
            prov_gr_wt = _safe_float(r.get('prov_gr_wt'))
            in_shop_wt = _safe_float(r.get('in_shop_wt'))
            in_transit_wt = _safe_float(r.get('in_transit_wt'))
            short_excess_wt = _safe_float(r.get('short_excess_wt'))
            percent = _safe_float(r.get('percent'))
            ordered_wt = _safe_float(r.get('ordered_wt'))

            if is_gt:
                has_grand = True
                row_fill = TOTAL_FILL
                row_font = BOLD_FONT
            else:
                row_font = BOLD_FONT if is_cat else DATA_FONT
                alt = (idx % 2 == 1)
                row_fill = ALT_FILL if alt else PatternFill(fill_type=None)
                if is_cat:
                    row_fill = CAT_FILL

                # Sum up totals
                if use_cats:
                    if is_cat:
                        if show_pcs:
                            totals['pcs'] += pcs
                        totals['prov_gr_wt'] += prov_gr_wt
                        totals['in_shop_wt'] += in_shop_wt
                        totals['in_transit_wt'] += in_transit_wt
                        totals['short_excess_wt'] += short_excess_wt
                        totals['ordered_wt'] += ordered_wt
                else:
                    if is_parent != 0:
                        if show_pcs:
                            totals['pcs'] += pcs
                        totals['prov_gr_wt'] += prov_gr_wt
                        totals['in_shop_wt'] += in_shop_wt
                        totals['in_transit_wt'] += in_transit_wt
                        totals['short_excess_wt'] += short_excess_wt
                        totals['ordered_wt'] += ordered_wt

            # Write cells
            col_idx = start_col
            
            # Label cell
            cell_lbl = ws.cell(row=curr_row, column=col_idx, value=label)
            cell_lbl.font = row_font
            cell_lbl.border = THIN_BORDER
            if row_fill.fill_type:
                cell_lbl.fill = row_fill
            if is_parent == 0:
                cell_lbl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            else:
                cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
            col_idx += 1

            # Values
            vals = []
            if show_pcs:
                vals.append((pcs, '#,##0', 'pcs'))
            vals.append((prov_gr_wt, '#,##0.000', 'prov_gr_wt'))
            vals.append((in_shop_wt, '#,##0.000', 'in_shop_wt'))
            vals.append((in_transit_wt, '#,##0.000', 'in_transit_wt'))
            vals.append((short_excess_wt, '#,##0.000', 'short_excess_wt'))
            vals.append((percent / 100.0, '0.0%', 'percent'))
            vals.append((ordered_wt, '#,##0.000', 'ordered_wt'))

            for val, num_fmt, key in vals:
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = row_font
                cell.border = THIN_BORDER
                if row_fill.fill_type:
                    cell.fill = row_fill
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = num_fmt
                
                # Text highlight for short/excess negative/positive
                if key == 'short_excess_wt' and not is_gt and not is_cat:
                    if val < 0:
                        cell.font = Font(name='Calibri', color='EF4444', size=9, bold=True) # Red
                    elif val > 0:
                        cell.font = Font(name='Calibri', color='16A34A', size=9) # Green
                        
                col_idx += 1

            curr_row += 1

        # Render Grand Total if not present
        if not has_grand and len(rows) > 0:
            row_fill = TOTAL_FILL
            row_font = BOLD_FONT
            
            # Label cell
            cell_lbl = ws.cell(row=curr_row, column=start_col, value='Grand Total')
            cell_lbl.font = row_font
            cell_lbl.border = THIN_BORDER
            cell_lbl.fill = row_fill
            cell_lbl.alignment = Alignment(horizontal='left', vertical='center')
            
            col_idx = start_col + 1
            
            # Values
            gt_percent = (totals['short_excess_wt'] * 100 / totals['prov_gr_wt']) if totals['prov_gr_wt'] != 0 else 0
            vals = []
            if show_pcs:
                vals.append((totals['pcs'], '#,##0'))
            vals.append((totals['prov_gr_wt'], '#,##0.000'))
            vals.append((totals['in_shop_wt'], '#,##0.000'))
            vals.append((totals['in_transit_wt'], '#,##0.000'))
            vals.append((totals['short_excess_wt'], '#,##0.000'))
            vals.append((gt_percent / 100.0, '0.0%'))
            vals.append((totals['ordered_wt'], '#,##0.000'))

            for val, num_fmt in vals:
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = row_font
                cell.border = THIN_BORDER
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.number_format = num_fmt
                col_idx += 1
                
            curr_row += 1

        return curr_row

    # ── Render Location Summary (at the top of Column 2) ─────────────────────────
    if 'Location Summary' in segments and segments['Location Summary']:
        col_2_row = render_excel_table(
            col_2_row, 9, 'Location Summary', segments['Location Summary'], show_pcs=True
        )
        col_2_row += 2

    # ── Grid Column 1 (Left Side: A-G) ──────────────────────────────────────────
    
    # 1. Purity Wise
    if 'Purity Wise' in segments:
        col_1_row = render_excel_table(
            col_1_row, 1, 'Purity Wise', segments['Purity Wise'], show_pcs=False
        )
        col_1_row += 2

    # 2. Classification Wise
    if 'Classification Wise' in segments:
        col_1_row = render_excel_table(
            col_1_row, 1, 'Classification Wise', segments['Classification Wise'],
            show_pcs=False, category_labels=['BRAND', 'GENERIC', 'LIFE STYLE']
        )
        col_1_row += 2

    # 3. Provision Mode Wise
    if 'Provision Mode Wise' in segments:
        col_1_row = render_excel_table(
            col_1_row, 1, 'Provision Mode Wise', segments['Provision Mode Wise'], show_pcs=False
        )
        col_1_row += 2

    # ── Grid Column 2 (Middle Side: I-P) ────────────────────────────────────────

    # 4. Section Wise
    if 'Section Wise' in segments:
        col_2_row = render_excel_table(
            col_2_row, 9, 'Section Wise', segments['Section Wise'], show_pcs=True
        )
        col_2_row += 2

    # 5. Section Details
    if 'Section Details' in segments:
        col_2_row = render_excel_table(
            col_2_row, 9, 'Section Details', segments['Section Details'], show_pcs=True
        )
        col_2_row += 2

    # ── Grid Column 3 (Right Side: R-X) ─────────────────────────────────────────

    # 6. Make Wise
    if 'Make Wise' in segments:
        col_3_row = render_excel_table(
            col_3_row, 18, 'Make Wise', segments['Make Wise'], show_pcs=False
        )
        col_3_row += 2

    # 7. Provision Type Wise
    if 'Provision Type Wise' in segments:
        col_3_row = render_excel_table(
            col_3_row, 18, 'Provision Type Wise', segments['Provision Type Wise'], show_pcs=False
        )
        col_3_row += 2

    # 8. Collection Wise
    if 'Collection Wise' in segments:
        col_3_row = render_excel_table(
            col_3_row, 18, 'Collection Wise', segments['Collection Wise'], show_pcs=False
        )
        col_3_row += 2

    # ── Filter Footer ────────────────────────────────────────────────────────────
    # Build a human-readable summary of active filters
    filter_label_map = [
        ('location', 'Location'),
        ('state', 'State'),
        ('purity', 'Purity'),
        ('classification', 'Classification'),
        ('make', 'Make'),
        ('collection', 'Collection'),
        ('section', 'Section'),
        ('prov_type', 'Provision Type'),
        ('provision_mode', 'Provision Mode'),
        ('branch_type', 'Branch Type'),
        ('branch_status', 'Branch Status'),
        ('business_head', 'Business Head'),
        ('search', 'Search'),
    ]

    active_parts = []
    for key, label in filter_label_map:
        val = filters.get(key, '')
        if isinstance(val, str):
            val = val.strip()
        if val:
            # Replace commas with ", " for readability
            display_val = str(val).replace(',', ', ')
            active_parts.append(f'{label}: {display_val}')

    # Sort info
    f_sort_by = filters.get('sort_by', '')
    f_sort_order = filters.get('sort_order', '')
    if f_sort_by and f_sort_order and f_sort_order != 'none':
        sort_label = f_sort_by.replace('_', ' ').title()
        active_parts.append(f'Sorted By: {sort_label} ({f_sort_order.upper()})')

    if active_parts:
        filter_text = 'Filters Applied:  ' + '  |  '.join(active_parts)
    else:
        filter_text = 'Filters Applied: None (All Data)'

    FOOTER_FONT = Font(name='Calibri', size=8, italic=True, color='6B7280')

    # Find the last used row across all 3 column cursors
    footer_row = max(col_1_row, col_2_row, col_3_row) + 2

    # Row 1: Filter summary (merged A:X)
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=24)
    filter_cell = ws.cell(row=footer_row, column=1, value=filter_text)
    filter_cell.font = FOOTER_FONT
    filter_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Row 2: Generated on timestamp
    footer_row += 1
    now_ist = datetime.now(IST)
    generated_text = f'Generated on: {now_ist.strftime("%d-%b-%Y %I:%M %p")} IST'
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=24)
    gen_cell = ws.cell(row=footer_row, column=1, value=generated_text)
    gen_cell.font = FOOTER_FONT
    gen_cell.alignment = Alignment(horizontal='left', vertical='center')

    # 9. Save
    timestamp = now_ist.strftime('%Y%m%d_%H%M%S')
    filename = f'location_physical_stock_status_{timestamp}.xlsx'
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)

    logger.info(f'Location Physical Stock Status export saved: {filepath}')
    return filename
